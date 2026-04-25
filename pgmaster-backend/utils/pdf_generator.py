"""
PDF receipt generation utilities for payment receipts.
"""

from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from django.core.files.base import ContentFile


def generate_payment_receipt(payment, pg_profile=None):
    """
    Generate PDF receipt for payment.
    
    Args:
        payment: Payment model instance
        pg_profile: PGProfile model instance
        
    Returns:
        BytesIO object containing PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=HexColor('#667eea'),
        spaceAfter=6,
        alignment=1  # Center
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=HexColor('#667eea'),
        spaceAfter=6,
        spaceBefore=12,
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=4,
    )
    
    # Header
    elements.append(Paragraph('PAYMENT RECEIPT', title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Property Info
    if pg_profile:
        elements.append(Paragraph(f'<b>{pg_profile.property_name}</b>', heading_style))
        elements.append(Paragraph(f'{pg_profile.address}', normal_style))
        elements.append(Paragraph(f'Phone: {pg_profile.contact_number}', normal_style))
        elements.append(Paragraph(f'Email: {pg_profile.owner.email}', normal_style))
    
    elements.append(Spacer(1, 0.2*inch))
    
    # Receipt Details Table
    receipt_data = [
        ['Receipt Number', f'RCP-{payment.id}-{datetime.now().strftime("%d%m%Y")}'],
        ['Receipt Date', datetime.now().strftime('%d %B %Y')],
        ['Receipt Time', datetime.now().strftime('%H:%M:%S')],
    ]
    
    receipt_table = Table(receipt_data, colWidths=[2*inch, 3.5*inch])
    receipt_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), HexColor('#f0f4ff')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(receipt_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Tenant Info
    elements.append(Paragraph('<b>TENANT INFORMATION</b>', heading_style))
    tenant_data = [
        ['Tenant Name', payment.tenant.tenant_name],
        ['Email', payment.tenant.email],
        ['Phone', payment.tenant.phone],
        ['Room', f'{payment.tenant.bed.room.room_number} (Floor {payment.tenant.bed.room.floor})'],
        ['Aadhar Number', payment.tenant.aadhar_number],
    ]
    
    tenant_table = Table(tenant_data, colWidths=[2*inch, 3.5*inch])
    tenant_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), HexColor('#f0f4ff')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(tenant_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Payment Details
    elements.append(Paragraph('<b>PAYMENT DETAILS</b>', heading_style))
    payment_data = [
        ['Description', 'Amount (₹)'],
        ['Monthly Rent', f'₹{payment.amount:,.2f}'],
        ['Late Fee', f'₹{payment.late_fee:,.2f}'],
        ['Extra Charges', f'₹{payment.extra_charges:,.2f}'],
        ['Total Amount Due', f'₹{payment.get_total_amount_due():,.2f}'],
        ['Amount Paid', f'₹{payment.paid_amount:,.2f}'],
        ['Balance', f'₹{payment.get_total_amount_due() - payment.paid_amount:,.2f}'],
    ]
    
    payment_table = Table(payment_data, colWidths=[2.5*inch, 3*inch])
    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('BACKGROUND', (0, 4), (-1, 4), HexColor('#fff3cd')),
        ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 6), (-1, 6), HexColor('#d1ecf1')),
        ('FONTNAME', (0, 6), (-1, 6), 'Helvetica-Bold'),
    ]))
    elements.append(payment_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Payment Method
    payment_methods = {
        'cash': 'Cash',
        'bank_transfer': 'Bank Transfer',
        'upi': 'UPI',
        'card': 'Credit/Debit Card',
        'cheque': 'Cheque',
    }
    
    elements.append(Paragraph(f'<b>Payment Method:</b> {payment_methods.get(payment.payment_method, payment.payment_method)}', normal_style))
    elements.append(Paragraph(f'<b>Transaction ID:</b> {payment.id}', normal_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Payment Status
    status_text = f'<b>Payment Status:</b> <font color="%s">{payment.get_status_display()}</font>' % (
        '#27ae60' if payment.status == 'paid' else
        '#f39c12' if payment.status == 'partial' else
        '#e74c3c'
    )
    elements.append(Paragraph(status_text, normal_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1,  # Center
    )
    
    elements.append(Paragraph('_' * 80, footer_style))
    elements.append(Spacer(1, 0.1*inch))
    elements.append(Paragraph('This is a system-generated receipt. No signature required.', footer_style))
    elements.append(Paragraph(f'Generated on {datetime.now().strftime("%d %B %Y at %H:%M:%S")}', footer_style))
    elements.append(Paragraph('© 2024 PGMaster. All rights reserved.', footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_occupancy_report_excel(pg_profile):
    """
    Generate Excel file for occupancy report.
    
    Args:
        pg_profile: PGProfile model instance
        
    Returns:
        BytesIO object containing Excel file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Occupancy Report'
    
    # Headers
    headers = ['Room', 'Floor', 'Type', 'Sharing', 'Bed', 'Status', 'Tenant Name', 'Phone']
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    row_num = 2
    rooms = pg_profile.rooms.all().prefetch_related('beds__tenant_set')
    
    for room in rooms:
        for bed in room.beds.all():
            tenant = bed.tenant_set.filter(status='active').first()
            
            worksheet.cell(row=row_num, column=1).value = room.room_number
            worksheet.cell(row=row_num, column=2).value = room.floor
            worksheet.cell(row=row_num, column=3).value = room.get_room_type_display()
            worksheet.cell(row=row_num, column=4).value = room.sharing_type
            worksheet.cell(row=row_num, column=5).value = bed.bed_number
            worksheet.cell(row=row_num, column=6).value = 'Occupied' if tenant else 'Vacant'
            worksheet.cell(row=row_num, column=7).value = tenant.tenant_name if tenant else '-'
            worksheet.cell(row=row_num, column=8).value = tenant.phone if tenant else '-'
            
            # Styling
            fill_color = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
            for col in range(1, len(headers) + 1):
                cell = worksheet.cell(row=row_num, column=col)
                cell.fill = fill_color
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            row_num += 1
    
    # Column widths
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 8
    worksheet.column_dimensions['C'].width = 12
    worksheet.column_dimensions['D'].width = 10
    worksheet.column_dimensions['E'].width = 8
    worksheet.column_dimensions['F'].width = 12
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 15
    
    # Save to BytesIO
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def generate_rent_report_excel(pg_profile):
    """
    Generate Excel file for pending rent report.
    
    Args:
        pg_profile: PGProfile model instance
        
    Returns:
        BytesIO object containing Excel file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from apps.payments.models import Payment
    from datetime import datetime, timedelta
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Rent Report'
    
    # Headers
    headers = ['Tenant Name', 'Room', 'Monthly Rent', 'Month', 'Status', 'Days Overdue', 'Amount Due', 'Phone']
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    row_num = 2
    payments = Payment.objects.filter(
        tenant__pg=pg_profile,
        status__in=['pending', 'partial', 'overdue']
    ).select_related('tenant')
    
    today = datetime.now().date()
    
    for payment in payments:
        days_overdue = (today - payment.due_date).days if payment.due_date < today else 0
        
        worksheet.cell(row=row_num, column=1).value = payment.tenant.tenant_name
        worksheet.cell(row=row_num, column=2).value = f'{payment.tenant.bed.room.room_number} (Floor {payment.tenant.bed.room.floor})'
        worksheet.cell(row=row_num, column=3).value = f'₹{payment.tenant.monthly_rent}'
        worksheet.cell(row=row_num, column=4).value = payment.month
        worksheet.cell(row=row_num, column=5).value = payment.get_status_display()
        worksheet.cell(row=row_num, column=6).value = days_overdue if days_overdue > 0 else 0
        worksheet.cell(row=row_num, column=7).value = f'₹{payment.get_total_amount_due()}'
        worksheet.cell(row=row_num, column=8).value = payment.tenant.phone
        
        # Styling
        fill_color = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid') if days_overdue > 0 else PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
        for col in range(1, len(headers) + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.fill = fill_color
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        row_num += 1
    
    # Column widths
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 12
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 12
    worksheet.column_dimensions['G'].width = 12
    worksheet.column_dimensions['H'].width = 15
    
    # Save to BytesIO
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
